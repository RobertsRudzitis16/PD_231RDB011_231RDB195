import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment 
import os

def find_popularity(p):
    p.rstrip()
    letter = p[-1]
    number = float(p[:-1])
    if letter == 'K':
        number *= 1000
    elif letter == 'M':
        number *= 1000000
    return number

print("Izvēlieties filmu saraksta kārtošanas metodi:")
print("1. Pēc reitinga")
print("2. Pēc gada")
print("3. Pēc popularitātes")
print("4. Pēc filmas garuma")
method = input("Ievadiet kārtošanas metodes numuru: ")
print()
if method in ['1','2','3','4']:
    method = int(method)
    order = ''
    while order not in ['1','2']:
        print("Izvēlieties filmu saraksta kārtošanas virzienu:")
        print("1. Augošā secībā")
        print("2. Dilstošā secībā")
        order = input("Ievadiet kārtošanas virziena numuru: ")

movies=[]
with open("movie_list.csv", "r", encoding="utf-8") as file:
    next(file)
    for line in file:
        film = line.split(',')
        movies.append(film[0]+' '+film[1])

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

url = "https://www.imdb.com/"
driver.get(url)
time.sleep(3)
find = driver.find_element(By.XPATH, '//*[@id="__next"]/div/div/div[2]/div/button[2]')
find.click()

if os.path.isfile("movies.xlsx"):
    os.remove("movies.xlsx")

wb = Workbook()
ws = wb.active
max_row=ws.max_row
ws["A1"] = "Movie title"
ws["B1"] = "Rating"
ws["C1"] = "Year"
ws["D1"] = "Popularity"
ws["E1"] = "Length"
ws["F1"] = "Actors"
ws["G1"] = "Director"
ws["H1"] = "Description"

data=[]

for line in movies:
    try:
        find = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.ID, "suggestion-search")))
        find.send_keys(line)
        find = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.ipc-metadata-list-summary-item__t")))
        find.click()

        name_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/h1/span')))
        rating_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[2]/div/div[1]/a/span/div/div[2]/div[1]/span[1]')))
        year_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/ul/li[1]')))
        popularity_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[2]/div/div[1]/a/span/div/div[2]/div[3]')))
        length_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[2]/div[1]/ul/li[3]')))
        actor1_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/section/div[2]/div/ul/li[3]/div/ul/li[1]/a')))
        actor2_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/section/div[2]/div/ul/li[3]/div/ul/li[2]/a')))
        actor3_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/section/div[2]/div/ul/li[3]/div/ul/li[3]/a')))
        director_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/section/div[2]/div/ul/li[1]/div/ul/li/a')))
        description_element = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/main/div/section[1]/section/div[3]/section/section/div[3]/div[2]/div[1]/section/p/span[2]')))
    
        name = name_element.text
        rating = rating_element.text
        year = year_element.text
        popularity = find_popularity(popularity_element.text)
        length = length_element.text
        actors = f"{actor1_element.text}, {actor2_element.text}, {actor3_element.text}"
        director = director_element.text
        description = description_element.text
        data.append([name, rating, f"'{year}", popularity, length, actors, director, description])
    except:
        continue
    
if method in [1,2,3,4]:
    if order == '1':
        data = sorted(data, key=lambda x: x[method])
    elif order == '2':
        data = sorted(data, key=lambda x: x[method], reverse=True)

for movie in data:
    ws.append(movie)

for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
    for cell in row:
        if i == 0:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
for column in ws.columns:
    max_length = max(len(str(cell.value)) for cell in column)
    ws.column_dimensions[column[0].column_letter].width = max_length + 2

wb.save("movies.xlsx")
wb.close